import json
from typing import Any

import openpyxl
from django.contrib import messages
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.db.models import QuerySet
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, render
from django.urls import resolve, reverse
from django.utils.translation import gettext_lazy as _
from django.views.generic import DetailView, View
from django_filters import rest_framework as django_filters
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework import filters as rest_filters
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.request import Request
from rest_framework.response import Response

from apps.core import filter_backends, mixins
from apps.core.inline import InlineFormsetFactory
from apps.core.models import Template
from apps.core.schemas import Action
from apps.core.utils import Deleter
from apps.fin.models import Compensation
from apps.hr.models import Employee

from .. import filters, forms, models, resources, serializers
from ..constants import vouchers as constants
from ..utils import (
    VoucherAuditor,
    VoucherDeleter,
    VoucherMigrator,
    VoucherUnMigrator,
)


class APIViewSet(
    mixins.APIMixin,
    mixins.BulkDeleteAPIMixin,
    viewsets.ModelViewSet,
):
    queryset = models.Voucher.objects.all().order_by("voucher_serial")
    serializer_class = serializers.VoucherSerializer
    filter_backends = [
        filter_backends.DjangoQLSearchFilter,
        django_filters.DjangoFilterBackend,
        rest_filters.OrderingFilter,
    ]
    filterset_class = filters.APIVoucherFilter
    ordering_fields = constants.ORDERING_FIELDS
    search_fields = constants.SEARCH_FIELDS
    deleter = VoucherDeleter

    @action(detail=True, methods=["post"], url_path="audit")
    def audit(self, request: Request, pk: int):
        instance = self.get_object()
        auditor = VoucherAuditor(request=request, obj=instance)
        auditor.action()

        if auditor.has_executed:
            return Response(status=status.HTTP_200_OK)
        else:
            return Response(
                {"details": auditor.get_message()},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"], url_path="migrate")
    def migrate(self, request: Request, pk: int):
        instance = self.get_object()
        migrator = VoucherMigrator(request=request, obj=instance)
        migrator.action()

        if migrator.has_executed:
            return Response(status=status.HTTP_200_OK)
        else:
            return Response(
                {"details": migrator.get_message()},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"], url_path="unmigrate")
    def unmigrate(self, request: Request, pk: int):
        instance = self.get_object()
        unmigrator = VoucherUnMigrator(request=request, obj=instance)
        unmigrator.action()

        if unmigrator.has_executed:
            return Response(status=status.HTTP_200_OK)
        else:
            return Response(
                {"details": unmigrator.get_message()},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["post"], url_path="bulk-audit")
    def bulk_audit(self, request: Request):
        ids = request.data.get("ids", [])
        qs: QuerySet = self.queryset.filter(pk__in=ids)

        if not qs.exists():
            return Response(
                {"details": _("no objects found")},
                status=status.HTTP_404_NOT_FOUND,
            )

        auditor = VoucherAuditor(request=self.request, queryset=qs)
        auditor.action()

        if auditor.has_executed:
            return Response(status=status.HTTP_200_OK)
        else:
            return Response(
                {"details": auditor.get_message()},
                status=status.HTTP_400_BAD_REQUEST,
            )


class ListView(PermissionRequiredMixin, mixins.ListMixin, View):
    permission_required = "trans.view_voucher"
    model = models.Voucher
    filter_class = filters.VoucherFilter
    resource_class = resources.VoucherResource
    ordering_fields = constants.ORDERING_FIELDS

    def get_actions(self) -> dict[str, Action]:
        return {
            "delete": Action(
                behavior=VoucherDeleter,
                template="components/blocks/modals/bulk-delete.html",
                permissions=("trans.delete_voucher",),
            ),
            "audit": Action(
                behavior=VoucherAuditor,
                template="components/trans/vouchers/modals/bulk-audit.html",
                permissions=("trans.audit_voucher",),
            ),
            "migrate": Action(
                behavior=VoucherMigrator,
                template="components/trans/vouchers/modals/bulk-migrate.html",
                permissions=("trans.migrate_voucher",),
                form_class=forms.VoucherBulkMigrateForm,
            ),
            "unmigrate": Action(
                behavior=VoucherUnMigrator,
                template="components/trans/vouchers/modals/bulk-unmigrate.html",
                permissions=("trans.unmigrate_voucher",),
            ),
        }


class DetailsView(PermissionRequiredMixin, mixins.DetailsMixin, DetailView):
    permission_required = "trans.view_voucher"
    model = models.Voucher


class CreateView(PermissionRequiredMixin, mixins.CreateMixin, View):
    permission_required = "trans.add_voucher"
    form_class = forms.VoucherForm

    def perform_create(self, form):
        obj: models.Voucher = form.save(commit=False)
        obj.created_by = self.request.user
        obj.updated_by = self.request.user
        obj.save()
        return obj


class VoucherTransactionInline(InlineFormsetFactory):
    model = models.VoucherTransaction
    form_class = forms.VoucherTransactionForm
    fields = ("employee", "compensation", "quantity", "value", "notes")
    deleter = Deleter

    @classmethod
    def get_queryset(cls, obj: models.VoucherTransaction):
        return obj.transactions.all().order_by("ordering", "id")


class UpdateView(PermissionRequiredMixin, mixins.UpdateMixin, View):
    permission_required = "trans.change_voucher"
    form_class = forms.VoucherForm
    inlines = (VoucherTransactionInline,)

    def can_access(self, request, obj: models.Voucher) -> bool:
        return not obj.is_migrated

    def cannot_access_message(self, request, obj: models.Voucher) -> str:
        return _("you cannot edit migrated voucher")

    def perform_update(self, form, formsets):
        obj = form.save(commit=False)
        for formset in formsets:
            last_order = 1
            for form in formset.forms:
                if form.is_valid() and form.cleaned_data:
                    item = form.save(commit=False)
                    item.ordering = form.cleaned_data.get("ORDER") or last_order
                    item.save()
                    last_order += 1
            formset.save()

        obj.updated_by = self.request.user
        obj.save()

        return obj


class DeleteView(PermissionRequiredMixin, mixins.BehaviorMixin, View):
    permission_required = "trans.delete_voucher"
    behavior = VoucherDeleter
    model = models.Voucher

    def can_access(self, request, obj: models.Voucher) -> bool:
        return not obj.is_migrated

    def cannot_access_message(self, request, obj: models.Voucher) -> str:
        return _("you cannot delete migrated voucher")


class AuditView(PermissionRequiredMixin, mixins.BehaviorMixin, View):
    permission_required = "trans.audit_voucher"
    behavior = VoucherAuditor
    model = models.Voucher
    modal_template_name = "components/trans/vouchers/modals/audit.html"


class MigrateView(PermissionRequiredMixin, mixins.BehaviorMixin, View):
    permission_required = "trans.migrate_voucher"
    behavior = VoucherMigrator
    model = models.Voucher
    modal_template_name = "components/trans/vouchers/modals/migrate.html"
    form_class = forms.VoucherMigrateForm


class UnMigrateView(PermissionRequiredMixin, mixins.BehaviorMixin, View):
    permission_required = "trans.unmigrate_voucher"
    behavior = VoucherUnMigrator
    model = models.Voucher
    modal_template_name = "components/trans/vouchers/modals/unmigrate.html"


class ExportToMSWordView(
    PermissionRequiredMixin,
    mixins.ExportToMSWordMixin,
    View,
):
    permission_required = "vouchers.export_voucher"

    def get_template(self):
        return Template.get_voucher_template()

    def get_filename(self):
        resolved = resolve(self.request.path_info)
        voucher = get_object_or_404(
            models.Voucher,
            slug=resolved.kwargs.get("slug"),
        )
        return voucher.voucher_serial

    def get_context_data(self):
        resolved = resolve(self.request.path_info)
        return {
            "voucher": models.Voucher.objects.get(
                slug=resolved.kwargs.get("slug"),
            ),
        }


class ImportView(View):
    model = models.Voucher
    form_class = forms.VoucherTransactionsImportForm

    def get(self, request: HttpRequest, slug: str, *args, **kwargs) -> HttpResponse:
        if request.GET.get("template"):
            return self.get_template_response()

        self.obj = get_object_or_404(self.model, slug=slug)

        template_name = self.get_template_name()
        context = self.get_context_data(obj=self.obj, table_id=self.get_table_id())

        return render(request, template_name, context)

    def post(self, request: HttpRequest, slug: str, *args, **kwargs) -> HttpResponse:
        self.obj = get_object_or_404(self.model, slug=slug)

        form = self.form_class(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return self.get_success_response(request, "ddd")
        else:
            return render(
                request=request,
                template_name=self.get_template_name(),
                context=self.get_context_data(obj=self.obj, form=form),
            )

    def get_template_response(self):
        filename = self.get_filename()
        response = HttpResponse(content_type="application/vnd.ms-excel")
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        wb = self.get_template_workbook()
        wb.save(response)

        return response

    def get_filename(self):
        return "Voucher Transactions template"

    def get_template_workbook(self):
        employees = Employee.objects.only("cost_center", "uuid", "fullname")
        compensations = Compensation.objects.only("name", "tax")

        wb = openpyxl.Workbook()

        wb.create_sheet("data")
        wb.create_sheet("employees")
        wb.create_sheet("compensations")

        employees_sheet: Worksheet = wb["employees"]
        employees_sheet.append(["fullname", "uuid"])
        for employee in employees:
            employees_sheet.append([employee.fullname, employee.uuid.hex])

        compensations_sheet: Worksheet = wb["compensations"]
        compensations_sheet.append(["name"])
        for compensation in compensations:
            compensations_sheet.append([compensation.name])

        wb["data"].append(
            [
                "employee",
                "employee uuid",
                "compensation",
                "quantity",
                "value",
                "notes",
            ]
        )

        return wb

    def get_codename_plural(self):
        return self.model._meta.codename_plural

    def get_app_label(self):
        return self.model._meta.app_label

    def get_index_url(self):
        app_label = self.get_app_label()
        codename_plural = self.get_codename_plural()
        return reverse(f"{app_label}:{codename_plural}:index")

    def get_table_id(self):
        return f"{self.get_codename_plural()}-table"

    def get_template_name(self):
        app_label = self.get_app_label()
        codename_plural = self.get_codename_plural()
        return f"components/{app_label}/{codename_plural}/modals/import.html"

    def get_context_data(self, **kwargs) -> dict[str, Any]:
        return {
            "form": self.form_class(),
            **kwargs,
        }

    def get_success_response(self, request: HttpRequest, message: str):
        response = HttpResponse()
        querystring = request.GET.urlencode() and f"?{request.GET.urlencode()}"
        messages.success(request, message)
        response["Hx-Location"] = json.dumps(
            {
                "path": self.get_index_url() + querystring,
                "target": f"#{self.get_table_id()}",
            }
        )
        response["HX-Trigger"] = "messages, hidemodal"
        return response

    def get_error_response(self, request: HttpRequest, message: str):
        response = HttpResponse()

        response["Hx-Retarget"] = "#no-content"
        response["HX-Reswap"] = "innerHTML"
        messages.error(request, message)
        response["HX-Trigger"] = "messages, hidemodal"

        return response
