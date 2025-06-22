from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl

from apps.edu.models import Degree, School, Specialization
from apps.geo.models import City, Nationality
from apps.hr.models import Employee, Mobile
from apps.org.models import CostCenter, JobSubtype, Position, Status


@dataclass
class EmployeeSchema:
    firstname: str = ""
    lastname: str = ""
    father_name: str = ""
    mother_name: str = ""
    birth_place: str = ""
    birth_date: str = ""
    national_id: str = ""
    card_id: str = ""
    civil_registry_office: str = ""
    registry_office_name: str = ""
    registry_office_id: str = ""
    gender: str = ""
    card_date: str = ""
    nationality: int = 1
    city: int = 1
    hire_date: str = ""
    cost_center: int = 1
    job_subtype: int = 1
    position: int = 1
    status: int = 1
    school: int = 1
    specialization: int = 1
    degree: int = 1
    religion: str = ""
    current_address: str = ""
    mobile: str = ""
    employee: Employee = field(init=False)

    def __post_init__(self):
        self.nationality = Nationality.objects.get(pk=self.nationality)
        self.city = City.objects.get(pk=self.city)
        self.cost_center = CostCenter.objects.get(pk=self.cost_center)
        self.job_subtype = JobSubtype.objects.get(pk=self.job_subtype)
        self.position = Position.objects.get(pk=self.position)
        self.status = Status.objects.get(pk=self.status)
        self.school = School.objects.get(pk=self.school)
        self.specialization = Specialization.objects.get(pk=self.specialization)
        self.degree = Degree.objects.get(pk=self.degree)

        if not self.hire_date:
            self.hire_date = datetime.now().date()

        self.birth_date = datetime.strftime(self.birth_date, "%Y-%m-%d")
        self.hire_date = datetime.strftime(self.hire_date, "%Y-%m-%d")

        if self.card_date:
            self.card_date = datetime.strftime(self.card_date, "%Y-%m-%d")
        else:
            self.card_date = None

        self.employee = Employee(
            firstname=self.firstname,
            lastname=self.lastname,
            father_name=self.father_name,
            mother_name=self.mother_name,
            birth_place=self.birth_place,
            birth_date=self.birth_date,
            national_id=self.national_id,
            card_id=self.card_id,
            civil_registry_office=self.civil_registry_office,
            registry_office_name=self.registry_office_name,
            registry_office_id=self.registry_office_id,
            gender=self.gender,
            card_date=self.card_date,
            nationality=self.nationality,
            city=self.city,
            hire_date=self.hire_date,
            cost_center=self.cost_center,
            job_subtype=self.job_subtype,
            position=self.position,
            status=self.status,
            school=self.school,
            specialization=self.specialization,
            degree=self.degree,
            current_address=self.current_address,
            religion=self.religion,
        )


def run():
    filepath = Path().resolve() / "scripts" / "employees_data.xlsx"

    if filepath.exists():
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            employee = EmployeeSchema(
                firstname=row[0].value,
                lastname=row[1].value,
                father_name=row[2].value,
                mother_name=row[3].value,
                birth_place=row[4].value,
                birth_date=row[5].value,
                national_id=row[6].value,
                card_id=row[7].value,
                civil_registry_office=row[8].value,
                registry_office_name=row[9].value,
                registry_office_id=row[10].value,
                gender=row[11].value,
                card_date=row[12].value,
                nationality=row[13].value,
                city=row[14].value,
                hire_date=row[15].value,
                cost_center=row[16].value,
                job_subtype=row[17].value,
                position=row[18].value,
                status=row[19].value,
                school=row[20].value,
                specialization=row[21].value,
                degree=row[22].value,
                religion=row[23].value,
                current_address=row[24].value,
                mobile=row[25].value,
            )
            print(f'{employee=}')
            employee.employee.save()
            if employee.mobile:
                Mobile(number=employee.mobile, employee=employee.employee).save()
